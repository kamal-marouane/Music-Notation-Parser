"""
Excel parser module for reading sheet music from Excel files.

This module provides utilities for reading and processing Excel files
containing sheet music notation.
"""

import os
import logging
from typing import Tuple, Optional, Dict, Any, Iterator
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from ..config import EXCEL_FILE_PATH, EXCEL_SHEET_NAME, TEMP_IMAGE_PATH
from ..utils.file_utils import ensure_file_exists

logger = logging.getLogger(__name__)


class ExcelReader:
    """
    Excel file reader for extracting music notation.
    
    This class reads cells from an Excel file containing musical notations
    and extracts both text and images.
    """
    
    def __init__(self, excel_path: str = EXCEL_FILE_PATH, sheet_name: str = EXCEL_SHEET_NAME):
        """
        Initialize the Excel reader.
        
        Args:
            excel_path: Path to the Excel file
            sheet_name: Name of the sheet to read
        """
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None
        self.image_loader = None
        self.current_cell = None
        self.current_text_iter = None
        self.current_row = 1
        self.current_col = 0
        
        # Initialize Excel reading
        self._initialize()
        
    def _initialize(self) -> None:
        """Initialize Excel workbook, sheet, and image loader."""
        ensure_file_exists(self.excel_path)
        
        try:
            self.workbook = openpyxl.load_workbook(self.excel_path)
            self.sheet = self.workbook[self.sheet_name]
            self.image_loader = SheetImageLoader(self.sheet)
            self._move_to_first_cell()
        except Exception as e:
            logger.error(f"Failed to initialize Excel reader: {e}")
            raise
            
    def _move_to_first_cell(self) -> None:
        """Move to the first cell (A1) in the sheet."""
        self.current_row = 1
        self.current_col = 0
        self._move_to_next_cell()
        
    def _get_column_letter(self, col_index: int) -> str:
        """
        Convert column index to Excel letter.
        
        Args:
            col_index: The 0-based column index
            
        Returns:
            The Excel column letter
        """
        col_letter = ""
        while col_index >= 0:
            col_index, remainder = divmod(col_index, 26)
            col_letter = chr(65 + remainder) + col_letter
            col_index -= 1
        return col_letter

    def _get_cell_reference(self) -> str:
        """
        Get the current cell reference (e.g., 'A1').
        
        Returns:
            Cell reference string
        """
        col_letter = self._get_column_letter(self.current_col)
        return f"{col_letter}{self.current_row}"
        
    def _move_to_next_cell(self) -> None:
        """Move to the next cell in the sheet."""
        self.current_col += 1
        cell_ref = self._get_cell_reference()
        self.current_cell = self.sheet[cell_ref]
        
        if self.current_cell.value:
            self.current_text_iter = iter(str(self.current_cell.value) + " ")
        else:
            self.current_text_iter = None
            
    def _move_to_next_row(self) -> None:
        """Move to the first column of the next row."""
        self.current_row += 1
        self.current_col = 0
        self._move_to_next_cell()
        
    def has_text_content(self) -> bool:
        """
        Check if current cell has text content.
        
        Returns:
            True if the cell has text content, False otherwise
        """
        return self.current_cell.value is not None
        
    def has_image_content(self) -> bool:
        """
        Check if current cell has an image.
        
        Returns:
            True if the cell has an image, False otherwise
        """
        cell_ref = self._get_cell_reference()
        try:
            return cell_ref in self.image_loader._images
        except Exception:
            return False
            
    def read_next_char(self) -> Optional[str]:
        """
        Read the next character from the current cell's text.
        
        Returns:
            The next character or None if no more characters
        """
        if not self.current_text_iter:
            return None
            
        try:
            return next(self.current_text_iter)
        except StopIteration:
            self._move_to_next_cell()
            return None
            
    def read_image(self) -> str:
        """
        Extract image from the current cell.
        
        Returns:
            Path to the extracted image
        """
        cell_ref = self._get_cell_reference()
        
        if not self.has_image_content():
            logger.warning(f"No image in cell {cell_ref}")
            return None
            
        try:
            image = self.image_loader.get(cell_ref)
            os.makedirs(TEMP_IMAGE_PATH, exist_ok=True)
            temp_image_path = os.path.join(TEMP_IMAGE_PATH, "temp_image.jpg")
            image.save(temp_image_path)
            self._move_to_next_cell()
            return temp_image_path
        except Exception as e:
            logger.error(f"Failed to extract image from cell {cell_ref}: {e}")
            self._move_to_next_cell()
            return None
            
    def move_to_next_empty_cell(self) -> None:
        """Move to the next cell that has no content."""
        while (self.has_text_content() or self.has_image_content()):
            self._move_to_next_cell()
            
            # If we reach the end of the row, move to the next row
            if self.current_col > 100:  # Safety limit
                self._move_to_next_row()